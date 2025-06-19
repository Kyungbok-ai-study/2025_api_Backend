"""
문제 및 정답 데이터 파싱 서비스 (Gemini 2.0 Flash 기반)

모든 파일 형식을 Gemini API로 통합 처리 - 모든 학과 지원
통합된 PDF 처리 및 배치 파싱 기능 포함
"""
import json
from typing import List, Dict, Any, Optional, Union, Callable
from datetime import datetime, timezone
import google.generativeai as genai
import os
from pathlib import Path
import base64
import logging
import re
import requests
import pandas as pd
from app.models.question import DifficultyLevel
from app.core.config import settings
from app.services.question_type_mapper import question_type_mapper
from app.services.evaluator_type_mapper import evaluator_type_mapper
# AI 문제 분석기 import 추가
from app.services.ai_question_analyzer import get_ai_analyzer

logger = logging.getLogger(__name__)

# Poppler 경로 설정 (PDF→이미지 변환용) - 클라우드 환경 최적화
_default_poppler = '/usr/bin'
POPPLER_PATH = os.getenv('POPPLER_PATH', _default_poppler)

# 학과 매핑
DEPARTMENT_MAPPING = {
    "물리치료학과": "물리치료",
    "작업치료학과": "작업치료", 
    "간호학과": "간호",
    "물리치료": "물리치료",
    "작업치료": "작업치료",
    "간호": "간호"
}

class QuestionParser:
    """gemini-2.0-flash-exp 기반 통합 파서 - 모든 학과 지원 + 통합 PDF 처리"""
    
    # 클래스 상수
    MAX_QUESTIONS = 22
    DEFAULT_YEAR = 2024
    DEFAULT_DIFFICULTY = "중"
    DEFAULT_DEPARTMENT = "물리치료학과"
    
    def __init__(self, api_key: Optional[str] = None):
        """
        Args:
            api_key: Gemini API 키 (None인 경우 환경변수에서 가져옴)
        """
        # 직접 API 키 설정
        self.api_key = api_key or "AIzaSyAU_5m68cNAMIBn7m1uQPrYKNFR0oPO3QA"
        if self.api_key:
            genai.configure(api_key=self.api_key)
            # Gemini 2.0 Flash 모델 사용 (정확한 모델명으로 수정)
            model_name = getattr(settings, 'GEMINI_MODEL_NAME', 'gemini-2.0-flash-exp')
            self.model = genai.GenerativeModel(model_name)
            logger.info(f"Gemini 모델 초기화 완료: {model_name}")
        else:
            self.model = None
            logger.warning("Gemini API 키가 설정되지 않았습니다.")
    
    def _apply_question_limit(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """22개 제한 로직 통합 메서드"""
        if not isinstance(data, list):
            return data
        
        # 문제번호가 22 이하인 것만 필터링 후 22개로 제한
        filtered_data = [item for item in data if item.get('question_number', 0) <= self.MAX_QUESTIONS]
        limited_data = filtered_data[:self.MAX_QUESTIONS]
        
        if len(filtered_data) > self.MAX_QUESTIONS:
            logger.info(f"📋 문제 제한 적용: {len(filtered_data)}개 → {len(limited_data)}개")
        
        return limited_data
    
    def _normalize_year(self, item: Dict[str, Any], fallback_year: Optional[int] = None) -> int:
        """연도 정규화 통합 메서드"""
        year = item.get('year')
        
        # 유효한 연도 체크
        if year and isinstance(year, int) and 2000 <= year <= 2030:
            return year
        
        # 문자열 연도 변환 시도
        if isinstance(year, str) and year.isdigit():
            year_int = int(year)
            if 2000 <= year_int <= 2030:
                return year_int
        
        # 폴백 연도 사용
        if fallback_year and 2000 <= fallback_year <= 2030:
            return fallback_year
        
        # 기본값
        return self.DEFAULT_YEAR
    
    def _get_normalized_department(self, department: str) -> str:
        """학과명 정규화 통합 메서드"""
        return DEPARTMENT_MAPPING.get(department, "물리치료")
    
    def detect_department_from_content(self, file_path: str, content_sample: str = "") -> str:
        """
        파일명과 내용으로부터 학과 자동 감지
        
        Args:
            file_path: 파일 경로
            content_sample: 파일 내용 샘플
            
        Returns:
            str: 감지된 학과명
        """
        file_name = Path(file_path).name.lower()
        content_lower = content_sample.lower()
        
        # 파일명 기반 감지
        if any(keyword in file_name for keyword in ['물치', '물리치료', 'pt', 'physical']):
            return "물리치료학과"
        elif any(keyword in file_name for keyword in ['작치', '작업치료', 'ot', 'occupational']):
            return "작업치료학과"
        elif any(keyword in file_name for keyword in ['간호', 'nursing', '너싱']):
            return "간호학과"
        
        # 내용 기반 감지
        if any(keyword in content_lower for keyword in ['물리치료', '재활의학', '운동치료', '전기치료']):
            return "물리치료학과"
        elif any(keyword in content_lower for keyword in ['작업치료', '인지치료', '감각통합', '보조기구']):
            return "작업치료학과"
        elif any(keyword in content_lower for keyword in ['간호학', '간호사', '환자간호', '임상간호']):
            return "간호학과"
        
        # 기본값 (파일명에서 추정)
        if '2.' in file_name or '3.' in file_name:
            return "작업치료학과"  # 작업치료 파일 패턴
        elif '1.' in file_name:
            return "물리치료학과"
        
        return "물리치료학과"  # 최종 기본값
    
    def parse_any_file(
        self, 
        file_path: str, 
        content_type: str = "auto", 
        department: str = "auto",
        progress_callback: Optional[Callable[[str, float], None]] = None
    ) -> Dict[str, Any]:
        """
        모든 파일 형식을 Gemini로 파싱 (분할 파싱 지원) - 모든 학과 지원
        
        Args:
            file_path: 파일 경로
            content_type: "questions", "answers", 또는 "auto" (자동 감지)
            department: 학과 정보 ("auto"인 경우 자동 감지)
            progress_callback: 진행률 콜백 함수 (message: str, progress: float)
            
        Returns:
            파싱된 데이터
        """
        if progress_callback:
            progress_callback("🚀 파싱 시작 중...", 0.0)
        
        if not self.model:
            logger.error("Gemini API가 초기화되지 않았습니다.")
            return {"type": content_type, "data": [], "error": "Gemini API not initialized"}

        if not os.path.exists(file_path):
            logger.error(f"파일이 존재하지 않습니다: {file_path}")
            return {"type": content_type, "data": [], "error": "File not found"}

        # 파일 크기 확인
        file_size = os.path.getsize(file_path)
        logger.info(f"📄 파일 크기: {file_size / (1024*1024):.2f} MB")
        
        if progress_callback:
            progress_callback(f"📄 파일 분석 중... ({file_size / (1024*1024):.2f} MB)", 5.0)

        # 학과 자동 감지
        if department == "auto":
            try:
                # 파일명으로 먼저 감지 시도
                detected_dept = self.detect_department_from_content(file_path)
                logger.info(f"🎯 학과 자동 감지: {detected_dept}")
                department = detected_dept
                
                if progress_callback:
                    progress_callback(f"🎯 학과 감지 완료: {department}", 10.0)
            except Exception as e:
                logger.warning(f"학과 자동 감지 실패: {e}, 기본값 사용")
                department = "물리치료학과"

        # DB 스키마 정보
        db_schema = f"""
우리 데이터베이스 구조 ({department} 전용):
Question 테이블:
- question_number: 문제 번호 (정수, 1~22까지만)
- content: 문제 내용 (텍스트)
- description: 문제 설명/지문 (문자열 배열, 예: ["- 설명1", "- 설명2"])
- options: {{"1": "선택지1", "2": "선택지2", ..., "5": "선택지5"}}
- correct_answer: 정답 (문자열, 예: "3")
- subject: 과목명 ({department} 관련)
- area_name: 영역이름 ({department} 전문 영역)
- difficulty: "하", "중", "상" 중 하나
- year: 연도 (정수)
중요: 22번 문제까지만 파싱하세요. 더 많은 문제가 있어도 22번까지만 처리하고 중단하세요.
"""

        try:
            # 파일 확장자에 따라 처리
            file_extension = Path(file_path).suffix.lower()

            if file_extension in ['.xlsx', '.xls']:
                if progress_callback:
                    progress_callback("📊 Excel 파일 처리 중...", 15.0)
                all_data = self._process_excel_file_chunked(file_path, content_type, db_schema, department, progress_callback)
            elif file_extension == '.pdf':
                if progress_callback:
                    progress_callback("📖 PDF 파일 처리 중...", 15.0)
                all_data = self._process_pdf_with_images(file_path, content_type, db_schema, progress_callback)
            else:
                if progress_callback:
                    progress_callback("📝 텍스트 파일 처리 중...", 15.0)
                all_data = self._process_text_file_chunked(file_path, content_type, db_schema, progress_callback)

            # 22개 제한 적용
            all_data = self._apply_question_limit(all_data)

            if progress_callback:
                progress_callback(f"📋 기본 파싱 완료: {len(all_data)}개 문제", 70.0)

            logger.info(f"분할 파싱 완료: {file_path}, 총 데이터 개수: {len(all_data)}")
            
            # 📊 3단계: AI 기반 문제 분석 (content_type이 questions인 경우)
            if content_type == "questions" and all_data:
                try:
                    if progress_callback:
                        progress_callback(f"🤖 AI 문제 분석 시작: {len(all_data)}개 문제", 75.0)
                    
                    logger.info(f"🤖 AI 문제 분석 시작: {len(all_data)}개 문제 ({department})")
                    
                    # AI 분석기 초기화
                    from app.services.ai_difficulty_analyzer import DifficultyAnalyzer
                    ai_analyzer = DifficultyAnalyzer()
                    
                    # 학과 정보 매핑
                    ai_department = DEPARTMENT_MAPPING.get(department, "물리치료")
                    
                    # 각 문제별 AI 분석
                    total_questions = len(all_data)
                    for idx, item in enumerate(all_data):
                        try:
                            content = item.get("content", "")
                            question_number = item.get("question_number", 1)
                            
                            # 진행률 업데이트
                            ai_progress = 75.0 + (idx / total_questions) * 20.0
                            if progress_callback:
                                progress_callback(f"🤖 문제 {question_number} AI 분석 중... ({idx+1}/{total_questions})", ai_progress)
                            
                            if content and content.strip():
                                # AI 분석
                                ai_result = ai_analyzer.analyze_question_auto(content, question_number, ai_department)
                                
                                if ai_result:
                                    # AI 분석 결과 추가
                                    item["ai_difficulty"] = ai_result.get("difficulty", "중")
                                    item["ai_question_type"] = ai_result.get("question_type", "객관식")
                                    item["ai_confidence"] = ai_result.get("confidence", "medium")
                                    item["ai_reasoning"] = ai_result.get("ai_reasoning", "AI 분석 완료")
                                    item["ai_analysis_complete"] = True
                                    item["updated_at"] = datetime.now().isoformat()
                                    
                                    # 기본 난이도 업데이트
                                    item["difficulty"] = ai_result.get("difficulty", "중")
                                    
                                    # 영역명은 AI 분석 결과 우선, 없으면 평가위원 데이터에서 조회
                                    area_name = ai_result.get("area_name")
                                    if not area_name or area_name == "일반":
                                        year = item.get("year", 2024)
                                        area_name = evaluator_type_mapper.get_area_name_for_question(
                                            department, year, question_number
                                        )
                                    item["area_name"] = area_name
                                    
                                    logger.info(f"✅ 문제 {question_number} AI 분석 완료: {ai_result.get('difficulty')} 난이도")
                                else:
                                    # AI 분석 실패 시 기본값
                                    item["ai_analysis_complete"] = False
                                    item["ai_reasoning"] = "AI 분석 불가능으로 기본 규칙 적용"
                                    
                                    # 영역명은 평가위원 데이터에서 조회
                                    year = item.get("year", 2024)
                                    area_name = evaluator_type_mapper.get_area_name_for_question(
                                        department, year, question_number
                                    )
                                    item["area_name"] = area_name
                                    
                                    logger.warning(f"⚠️ 문제 {question_number} AI 분석 실패")
                            else:
                                logger.warning(f"⚠️ 문제 {question_number} content 없음으로 AI 분석 건너뜀")
                                
                                # 영역명은 평가위원 데이터에서 조회
                                year = item.get("year", 2024)
                                area_name = evaluator_type_mapper.get_area_name_for_question(
                                    department, year, question_number
                                )
                                item["area_name"] = area_name
                                
                        except Exception as e:
                            logger.error(f"❌ 문제 {item.get('question_number')} AI 분석 오류: {e}")
                            item["ai_analysis_complete"] = False
                            item["ai_reasoning"] = f"AI 분석 오류: {str(e)}"
                            
                            # 영역명은 평가위원 데이터에서 조회
                            year = item.get("year", 2024)
                            question_number = item.get("question_number", 1)
                            area_name = evaluator_type_mapper.get_area_name_for_question(
                                department, year, question_number
                            )
                            item["area_name"] = area_name
                    
                    if progress_callback:
                        progress_callback(f"🎯 AI 분석 완료: {len(all_data)}개 문제 처리됨", 95.0)
                    
                    logger.info(f"🎯 AI 분석 완료: {len(all_data)}개 문제 처리됨")
                    
                except Exception as e:
                    logger.error(f"❌ AI 분석 전체 실패: {e}")
                    if progress_callback:
                        progress_callback(f"⚠️ AI 분석 실패, 기본 파싱 결과 사용", 95.0)
                    # AI 분석 실패해도 파싱은 계속 진행
            
            if progress_callback:
                progress_callback("✅ 파싱 완료!", 100.0)
            
            return {
                "type": content_type if content_type != "auto" else "questions", 
                "data": all_data,
                "department": department,
                "total_questions": len(all_data)
            }

        except Exception as e:
            logger.error(f"분할 파싱 오류 ({file_path}): {e}")
            if progress_callback:
                progress_callback(f"❌ 파싱 오류: {str(e)}", 0.0)
            return {"type": content_type, "data": [], "error": str(e)}
    
    def _generate_prompt(self, file_path: str, content_type: str, db_schema: str) -> str:
        """프롬프트 생성"""
        file_name = Path(file_path).name
        
        if content_type == "auto":
            return f"""다음 파일을 분석해주세요.

파일명: {file_name}

이 파일이 문제 데이터인지 정답 데이터인지 자동으로 판단하고,
아래 데이터베이스 구조에 맞게 JSON으로 변환해주세요.

{db_schema}

반환 형식:
{{
    "type": "questions" 또는 "answers",
    "data": [
        // 위 스키마에 맞는 객체들
    ]
}}

주의사항:
- 문제번호는 반드시 정수로, 22번까지만 처리
- 선택지 번호는 "1", "2", "3", "4", "5" 문자열로
- 난이도: "하", "중", "상" 중 하나로 표시
- 연도는 파일명이나 내용에서 추출
- 없는 필드는 null로 표시
- JSON 형식으로만 응답해주세요"""
            
        elif content_type == "questions":
            return f"""이 파일은 시험 문제입니다.
            
{db_schema}

위 Question 스키마에 맞게 모든 문제를 JSON 배열로 변환해주세요.
선택지가 ①②③④⑤로 되어있다면 "1", "2", "3", "4", "5"로 변환하세요.

중요 제한사항:
- 22번 문제까지만 파싱하세요. 더 많은 문제가 있어도 22번에서 중단하세요.
- 문제번호가 22를 초과하면 무시하세요.

주의사항:
- 문제에 보충 설명이나 지문이 있으면 description 필드에 배열로 저장하세요
- description은 문제를 풀기 위한 추가 정보나 조건들을 담습니다
- 예: ["- 몸에 널리 분포하며, 몸의 구조를 이룸", "- 세포나 기관 사이 틈을 메우고, 기관을 지지·보호함"]
- JSON 형식으로만 응답해주세요"""
        else:  # answers
            return f"""이 파일은 시험 정답지입니다. 각 문제 번호와 해당 정답을 정확히 추출해주세요.

{db_schema}

답안지에서 다음 정보를 찾아 JSON 배열로 변환해주세요:

**추출 대상:**
- question_number: 문제 번호 (1, 2, 3, ... 형태의 숫자)
- correct_answer: 정답 (1, 2, 3, 4, 5 중 하나)
- subject: 과목명 (이미지/텍스트에서 식별 가능하면)
- area_name: 영역명 (이미지/텍스트에서 식별 가능하면)
- difficulty: 난이도 (이미지/텍스트에서 식별 가능하면, 없으면 null)
- year: 연도 (파일명이나 내용에서 추출)

**답안지 인식 패턴:**
- "1번: ②", "2번: ①", "3번: ⑤" 형태
- "1. ②", "2. ①", "3. ⑤" 형태  
- "문제 1: 정답 2", "문제 2: 정답 1" 형태
- 표 형태로 된 답안 (문제번호 | 정답)
- ①②③④⑤ 기호는 1,2,3,4,5로 변환

**출력 예시:**
[
  {
    "question_number": 1,
    "correct_answer": "2",
    "year": 2020
  },
  {
    "question_number": 2,
    "correct_answer": "1",
    "year": 2020
  }
]

**중요 제한사항:**
- 22번 문제까지만 파싱하세요
- 정답은 반드시 "1", "2", "3", "4", "5" 중 하나의 문자열로
- 문제 번호가 명확하지 않은 경우 순서대로 1,2,3... 배정
- JSON 형식으로만 응답하세요"""
    
    async def _process_excel_file_chunked(
        self, 
        file_path: str, 
        content_type: str, 
        db_schema: str, 
        department: str = "물리치료학과",
        progress_callback: Optional[Callable[[str, float], None]] = None
    ) -> List[Dict[str, Any]]:
        """Excel 파일 분할 처리 (openpyxl 사용) - 모든 학과 지원"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 라이브러리가 설치되지 않았습니다.")
            raise ImportError("openpyxl이 필요합니다. pip install openpyxl로 설치하세요.")
        
        all_data = []
        
        if progress_callback:
            progress_callback(f"📊 Excel 파일 로드 중... ({department})", 20.0)
        
        # 📊 1단계: 문제 유형 매핑 데이터 생성 (content_type이 questions인 경우)
        if content_type == "questions":
            try:
                # 교수명 추출 (파일명에서)
                professor_name = self._extract_professor_from_filename(file_path)
                
                if progress_callback:
                    progress_callback(f"🎯 문제 유형 자동 배정 중... ({professor_name})", 25.0)
                
                # 문제 유형 매핑 처리
                logger.info(f"🎯 문제 유형 자동 배정 시작: {professor_name} ({department})")
                type_result = await question_type_mapper.process_excel_for_question_types(
                    file_path, professor_name, department
                )
                
                if type_result.get("success"):
                    logger.info(f"✅ 문제 유형 매핑 완료: {type_result['total_questions']}개 문제")
                    self.question_type_file_key = type_result["file_key"]
                else:
                    logger.warning(f"⚠️ 문제 유형 매핑 실패: {type_result.get('error', '알 수 없는 오류')}")
                    self.question_type_file_key = None
                    
            except Exception as e:
                logger.warning(f"⚠️ 문제 유형 매핑 과정에서 오류: {e}")
                self.question_type_file_key = None
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            logger.info(f"Excel 파일 시트 목록: {workbook.sheetnames}")
            
            if progress_callback:
                progress_callback(f"📊 시트 분석 중: {len(workbook.sheetnames)}개 시트", 30.0)
            
            total_sheets = len(workbook.sheetnames)
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                logger.info(f"시트 '{sheet_name}' 처리 중... ({sheet_idx+1}/{total_sheets})")
                
                sheet_progress = 30.0 + (sheet_idx / total_sheets) * 30.0
                if progress_callback:
                    progress_callback(f"📄 시트 '{sheet_name}' 처리 중... ({sheet_idx+1}/{total_sheets})", sheet_progress)
                
                # 시트 데이터를 텍스트로 변환
                sheet_data = []
                row_count = 0
                
                for row in worksheet.iter_rows(values_only=True):
                    if row_count >= 100:  # 시트당 최대 100행
                        break
                    if any(cell for cell in row):  # 빈 행이 아닌 경우만
                        sheet_data.append('\t'.join([str(cell) if cell is not None else '' for cell in row]))
                        row_count += 1
                
                if sheet_data:
                    sheet_text = '\n'.join(sheet_data)
                    
                    # Gemini로 구조화 요청
                    prompt = self._generate_prompt(f"{file_path} (시트: {sheet_name})", content_type, db_schema)
                    structured_prompt = f"""
다음은 Excel 시트 '{sheet_name}'의 데이터입니다 ({department} 전용).
이 데이터를 분석하여 구조화된 JSON으로 변환해주세요.

{prompt}

Excel 데이터:
{sheet_text}

중요: 22번 문제까지만 처리하세요.
학과: {department}
"""
                    
                    try:
                        response = self.model.generate_content([structured_prompt])
                        if response and response.text:
                            sheet_results = self._parse_gemini_response(response.text, content_type)
                            sheet_data_parsed = sheet_results.get("data", [])
                            
                            # 22번 제한 적용
                            sheet_data_parsed = self._apply_question_limit(sheet_data_parsed)
                            
                            if sheet_data_parsed:
                                # year 보정: Gemini가 year를 못 뽑았거나 0/None이면 시트명에서 추출
                                year_in_sheet = None
                                match = re.search(r'(20\d{2})', sheet_name)
                                if match:
                                    year_in_sheet = int(match.group(1))
                                else:
                                    year_in_sheet = 2020  # 기본값
                                
                                # 📊 2단계: 각 문제에 유형 정보 추가
                                for item in sheet_data_parsed:
                                    if not item.get('year') or item.get('year') in [0, None, '']:
                                        item['year'] = year_in_sheet
                                    
                                    # 학과 정보 추가
                                    item['department'] = department
                                    
                                    # 문제 유형 자동 배정 (questions인 경우만)
                                    if content_type == "questions" and hasattr(self, 'question_type_file_key'):
                                        question_content = item.get('content', '')
                                        question_number = item.get('question_number')
                                        
                                        # 문제 유형 조회 (기존 규칙 기반)
                                        question_type = question_type_mapper.get_question_type_for_question(
                                            question_content, 
                                            self.question_type_file_key, 
                                            question_number
                                        )
                                        
                                        # 문제 유형 정보 추가
                                        item['question_type'] = question_type
                                        item['type_name'] = question_type_mapper.question_types.get(
                                            question_type, {}
                                        ).get('name', question_type)
                                        
                                        logger.debug(f"   문제 {question_number}: {question_type} ({item['type_name']})")
                                
                                logger.info(f"시트 '{sheet_name}': {len(sheet_data_parsed)}개 항목 파싱 성공 (연도 보정: {year_in_sheet})")
                                all_data.extend(sheet_data_parsed)
                            else:
                                logger.warning(f"시트 '{sheet_name}': 파싱된 데이터 없음")
                    except Exception as e:
                        logger.error(f"시트 '{sheet_name}' 파싱 실패: {e}")
                        continue
                        
                # 22개 달성하면 중단
                if len(all_data) >= self.MAX_QUESTIONS:
                    all_data = self._apply_question_limit(all_data)
                    break
            
            workbook.close()
            
            if progress_callback:
                progress_callback(f"📊 Excel 파싱 완료: {len(all_data)}개 문제", 60.0)
            
            logger.info(f"Excel 파일 처리 완료: 총 {len(all_data)}개 항목")
            return all_data
            
        except Exception as e:
            logger.error(f"Excel 파일 읽기 오류: {e}")
            raise
    
    def _extract_professor_from_filename(self, file_path: str) -> str:
        """파일명에서 교수명 추출"""
        try:
            filename = Path(file_path).name
            # "2. 신장훈_작치_마스터코딩지.xlsx" -> "신장훈"
            if "_" in filename:
                parts = filename.split("_")
                if len(parts) >= 2:
                    name_part = parts[0].replace("2. ", "").strip()
                    return name_part
            
            # 기본값: 파일명에서 확장자 제거
            return Path(file_path).stem
            
        except Exception as e:
            logger.warning(f"교수명 추출 실패: {e}")
            return "Unknown"
    
    def _process_pdf_with_images(
        self, 
        file_path: str, 
        content_type: str, 
        db_schema: str,
        progress_callback: Optional[Callable[[str, float], None]] = None
    ) -> List[Dict[str, Any]]:
        """PDF 파일을 이미지로 변환하여 Gemini로 처리 (통합 PDF 처리 사용)"""
        all_questions = []
        
        try:
            if progress_callback:
                progress_callback("📖 PDF → 이미지 변환 중...", 20.0)
            
            # 통합된 PDF → 이미지 변환 사용
            logger.info("PDF를 이미지로 변환 중...")
            page_images_base64 = self._convert_pdf_to_images_unified(file_path, max_pages=20)
            
            if not page_images_base64:
                # 이미지 변환 실패시 텍스트 추출 시도
                logger.warning("이미지 변환 실패, 텍스트 추출 시도...")
                text_content = self._extract_pdf_text_fallback(file_path)
                if text_content:
                    return self._process_text_chunks(text_content, content_type, db_schema, progress_callback)
                else:
                    raise Exception("PDF 처리 실패: 이미지 변환과 텍스트 추출 모두 실패")
            
            logger.info(f"총 {len(page_images_base64)}개 페이지 이미지 생성됨")
            
            if progress_callback:
                progress_callback(f"📄 {len(page_images_base64)}개 페이지 이미지 생성 완료", 40.0)
            
            # 파일 타입별 프롬프트 생성
            if content_type == "answers":
                # 답안지 전용 강화 프롬프트
                gemini_prompt = f"""
이 이미지는 시험 정답지/답안지입니다. 이미지에서 문제 번호와 정답을 찾아 추출해주세요.

**중요: 이미지에 있는 모든 숫자와 선택지 기호를 꼼꼼히 살펴보세요!**

**찾아야 할 패턴들:**
1. "1번 ②", "2번 ①", "3번 ④" 
2. "1. ②", "2. ①", "3. ④"
3. "문제1 정답②", "문제2 정답①"
4. "1번문제: ②", "2번문제: ①"
5. "1-②", "2-①", "3-④"
6. 표 형태: | 1 | ② | 또는 | 문제1 | 정답② |
7. "정답: 1②, 2①, 3④..."
8. 세로로 나열된 형태도 찾아보세요

**선택지 변환 규칙:**
- ① → "1"
- ② → "2" 
- ③ → "3"
- ④ → "4"
- ⑤ → "5"
- 1번 → "1"
- 2번 → "2"
- A → "1", B → "2", C → "3", D → "4", E → "5"

**출력 형식 (JSON 배열만):**
[
  {{"question_number": 1, "correct_answer": "2", "year": 2021}},
  {{"question_number": 2, "correct_answer": "1", "year": 2021}},
  {{"question_number": 3, "correct_answer": "4", "year": 2021}}
]

**주의사항:**
- 이미지에서 보이는 모든 문제-정답 쌍을 찾으세요
- 22번까지만 추출하세요
- 문제번호가 명확하지 않으면 순서대로 1,2,3... 배정하세요
- 정답이 보이지 않는 문제는 제외하세요
- 반드시 JSON 배열 형식으로만 응답하세요

이미지를 자세히 보고 정답지의 모든 정보를 놓치지 마세요!
"""
            else:
                # 문제지 전용 프롬프트
                gemini_prompt = f"""
이 이미지를 분석하여 시험 문제를 찾아 구조화된 JSON으로 변환해주세요.

{db_schema}

위 스키마에 맞게 각 문제를 추출하세요:
- question_number: 문제 번호 (1~22만)
- content: 문제 내용
- description: 문제 설명/지문이 있다면 배열로
- options: 선택지 (①②③④⑤ → "1","2","3","4","5")
- year: 연도 (이미지에서 추출 가능하면)

JSON 배열로만 응답하세요. 22번 문제까지만 처리하세요.
"""
            
            # 💀 CRITICAL: 모든 페이지에서 문제 추출 (22개까지)
            question_numbers_found = set()
            total_pages = len(page_images_base64)
            
            for page_num, page_image_base64 in enumerate(page_images_base64, 1):
                # base64를 Gemini용 이미지 객체로 변환
                import io
                import base64
                from PIL import Image
                
                try:
                    image_data = base64.b64decode(page_image_base64)
                    page_image = Image.open(io.BytesIO(image_data))
                except Exception as e:
                    logger.error(f"페이지 {page_num} 이미지 디코딩 실패: {e}")
                    continue
                page_progress = 40.0 + (page_num / total_pages) * 50.0
                if progress_callback:
                    progress_callback(f"📖 페이지 {page_num}/{total_pages} 이미지 분석 중...", page_progress)
                
                logger.info(f"📖 페이지 {page_num}/{total_pages} 이미지 분석 중...")
                
                try:
                    # Gemini 분석
                    response = self.model.generate_content([gemini_prompt, page_image])
                    
                    if response and response.text:
                        try:
                            # 🔍 답안지인 경우 상세 디버깅
                            if content_type == "answers":
                                logger.info(f"🔍 페이지 {page_num} Gemini 원본 응답: {response.text[:500]}...")
                            
                            # 응답 파싱
                            page_result = self._parse_gemini_response(response.text, content_type)
                            page_questions = page_result.get("data", [])
                            
                            # 🔍 답안지인 경우 파싱 결과 로깅
                            if content_type == "answers":
                                logger.info(f"🔍 페이지 {page_num} 파싱 결과: {len(page_questions)}개 항목")
                                for i, q in enumerate(page_questions[:3]):  # 처음 3개만 로깅
                                    logger.info(f"    항목 {i+1}: {q}")
                            
                            # 유효한 데이터 필터링 (답안지는 다른 검증 기준)
                            valid_page_data = []
                            
                            if content_type == "answers":
                                # 답안지: question_number와 correct_answer만 있으면 유효
                                for q in page_questions:
                                    q_num = q.get('question_number', 0)
                                    answer = q.get('correct_answer', '')
                                    
                                    if (1 <= q_num <= 22 and 
                                        answer and answer.strip() and answer in ["1", "2", "3", "4", "5"] and
                                        q_num not in question_numbers_found):
                                        
                                        valid_page_data.append(q)
                                        question_numbers_found.add(q_num)
                                        logger.info(f"✅ 정답 {q_num}: {answer}")
                            else:
                                # 문제지: 기존 검증 방식
                                for q in page_questions:
                                    q_num = q.get('question_number', 0)
                                    content = q.get('content', '')
                                    
                                    if (1 <= q_num <= 22 and 
                                        content and content.strip() and content != "null" and
                                        q_num not in question_numbers_found):
                                        
                                        valid_page_data.append(q)
                                        question_numbers_found.add(q_num)
                                        logger.info(f"✅ 문제 {q_num} 추출 성공")
                            
                            if valid_page_data:
                                all_questions.extend(valid_page_data)
                                logger.info(f"📄 페이지 {page_num}: {len(valid_page_data)}개 신규 데이터 추가 (총 {len(all_questions)}개)")
                            else:
                                logger.warning(f"⚠️ 페이지 {page_num}: 유효한 데이터 없음")
                                if content_type == "answers" and page_questions:
                                    logger.warning(f"    원본 데이터: {page_questions}")
                                
                        except Exception as e:
                            logger.error(f"❌ 페이지 {page_num} 파싱 실패: {e}")
                            if content_type == "answers":
                                logger.error(f"    원본 응답: {response.text[:300]}...")
                            continue
                    else:
                        logger.warning(f"⚠️ 페이지 {page_num}: Gemini 응답 없음")
                    
                    # 22개 달성 확인
                    if len(question_numbers_found) >= 22:
                        logger.info(f"🎯 22문제 달성! 더 이상 처리하지 않음")
                        if progress_callback:
                            progress_callback("🎯 22문제 달성! 파싱 완료", 90.0)
                        break
                    
                except Exception as e:
                    logger.error(f"❌ 페이지 {page_num} 전체 실패: {e}")
                    continue
            
            # 최종 22개 제한 적용
            all_questions = all_questions[:22]
            
            if progress_callback:
                progress_callback(f"📖 PDF 이미지 분석 완료: {len(all_questions)}개 문제", 90.0)
            
            logger.info(f"PDF 이미지 분석 완료: 총 {len(all_questions)}개 문제")
            return all_questions
            
        except Exception as e:
            logger.error(f"PDF 이미지 처리 실패: {e}")
            if progress_callback:
                progress_callback(f"❌ PDF 처리 실패: {str(e)}", 0.0)
            raise

    def _process_text_file_chunked(self, file_path: str, content_type: str, db_schema: str, progress_callback: Optional[Callable[[str, float], None]] = None) -> List[Dict[str, Any]]:
        """텍스트 파일 분할 처리"""
        encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError(f"파일 인코딩을 감지할 수 없습니다: {file_path}")
        
        return self._process_text_chunks(content, content_type, db_schema, progress_callback)

    def _process_text_chunks(self, content: str, content_type: str, db_schema: str, progress_callback: Optional[Callable[[str, float], None]] = None) -> List[Dict[str, Any]]:
        """텍스트 내용을 청크로 분할 처리"""
        all_data = []
        chunk_size = 15000  # 문자 단위 청크 크기
        
        # 텍스트가 작은 경우 한번에 처리
        if len(content) <= chunk_size:
            try:
                prompt = self._generate_prompt("text_content", content_type, db_schema)
                response = self.model.generate_content([prompt, f"텍스트 내용:\n{content}"])
                result = self._parse_gemini_response(response.text, content_type).get("data", [])
                # 22번 제한 적용
                result = [item for item in result if item.get('question_number', 0) <= 22][:22]
                return result
            except Exception as e:
                logger.error(f"텍스트 처리 오류: {e}")
                return []
        
        # 큰 텍스트는 청크로 분할
        for i in range(0, len(content), chunk_size):
            chunk = content[i:i + chunk_size]
            
            try:
                prompt = self._generate_prompt("text_content", content_type, db_schema)
                response = self.model.generate_content([prompt, f"텍스트 청크:\n{chunk}"])
                chunk_data = self._parse_gemini_response(response.text, content_type).get("data", [])
                
                # 22번 제한 적용
                chunk_data = self._apply_question_limit(chunk_data)
                
                all_data.extend(chunk_data)
                logger.info(f"텍스트 청크 처리 완료: {len(chunk_data)}개 데이터")
                
                # 22개 달성하면 중단
                if len(all_data) >= self.MAX_QUESTIONS:
                    all_data = self._apply_question_limit(all_data)
                    break
                    
            except Exception as e:
                logger.warning(f"텍스트 청크 처리 실패: {e}")
                continue
        
        return all_data

    # 중복된 JSON 정리 메소드들은 통합 static 메소드로 대체됨
    # _clean_json_text_unified 및 _aggressive_json_clean_unified 사용
    
    def _parse_gemini_response(self, response_text: str, content_type: str) -> Dict[str, Any]:
        """Gemini 응답 파싱 (통합 유틸리티 사용)"""
        
        try:
            # 통합 AI JSON 파서 사용
            result = self.parse_ai_json_response(
                response_text,
                fallback_data={"error": "파싱 실패", "data": [], "type": content_type}
            )
            
            # 에러 응답 확인
            if "error" in result:
                logger.error(f"JSON 파싱 실패: {result['error']}")
                return {"type": content_type, "data": []}
            
            # 자동 감지 모드인 경우
            if content_type == "auto" and isinstance(result, dict) and "type" in result:
                data = result.get("data", [])
                # 22번 제한 적용
                data = self._apply_question_limit(data)
                return {
                    "type": result["type"],
                    "data": data
                }
            else:
                # 지정된 타입인 경우 - 데이터 정규화
                if isinstance(result, list):
                    data = result
                elif isinstance(result, dict):
                    data = result.get("data", [result] if result else [])
                else:
                    data = []
                
                # 22번 제한 적용
                data = self._apply_question_limit(data)
                return {
                    "type": content_type,
                    "data": data
                }
                
        except Exception as e:
            logger.error(f"❌ Gemini 응답 파싱 중 예외: {e}")
            return {"type": content_type, "data": []}
    
    def match_questions_with_answers(
        self, 
        questions: List[Dict[str, Any]], 
        answers: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        문제와 정답 매칭 (22개 제한)
        
        정답이 없는 문제도 포함시키되 correct_answer를 빈 값으로 설정
        22번 문제까지만 처리합니다.
        """
        # 입력 데이터에 22개 제한 적용
        questions = [q for q in questions if q.get('question_number', 0) <= self.MAX_QUESTIONS][:self.MAX_QUESTIONS]
        answers = [a for a in answers if a.get('question_number', 0) <= self.MAX_QUESTIONS]

        # 정답을 문제번호로 인덱싱
        answer_map = {}
        for ans in answers:
            q_num = ans.get("question_number")
            if q_num is not None and q_num <= self.MAX_QUESTIONS:  # 22번까지만
                answer_map[str(q_num)] = ans

        matched_data = []
        matched_count = 0

        # 정답이 있는 문제번호 범위 확인
        if answer_map:
            available_answer_numbers = set(answer_map.keys())
            logger.info(f"사용 가능한 정답: {len(available_answer_numbers)}개 문제 ({min(available_answer_numbers) if available_answer_numbers else 'N/A'} ~ {max(available_answer_numbers) if available_answer_numbers else 'N/A'}번)")
        else:
            logger.warning("정답 데이터가 없습니다. 모든 문제를 정답 없이 포함합니다.")
            available_answer_numbers = set()

        for question in questions:
            q_num = question.get("question_number")
            if q_num is None:
                logger.warning(f"문제번호가 없는 문제 건너뛰기: {question.get('content', '')[:50]}...")
                continue

            q_num_str = str(q_num)

            # 기본 문제 데이터 설정
            matched_item = {
                **question,
                "correct_answer": "",
                "answer_source": "no_answer"
            }

            # 정답이 있는 경우 병합
            if q_num_str in answer_map:
                answer_data = answer_map[q_num_str]
                matched_item.update({
                    "correct_answer": answer_data.get("correct_answer") or answer_data.get("answer", ""),
                    "subject": answer_data.get("subject", question.get("subject", "")),
                    "area_name": answer_data.get("area_name", question.get("area_name", "")),
                    "difficulty": answer_data.get("difficulty", question.get("difficulty", "중")),
                    "year": answer_data.get("year", question.get("year")),
                    "answer_source": "matched"
                })
                matched_count += 1
                logger.debug(f"✅ 문제 {q_num}: 정답 매칭 완료")
            else:
                logger.debug(f"⚠️ 문제 {q_num}: 정답 없음, 빈 값으로 설정")

            # 기본 필수 필드 검증 (content만 확인)
            if matched_item.get("content") and matched_item.get("content").strip():
                matched_data.append(matched_item)
            else:
                logger.warning(f"문제 {q_num}: content가 없어 제외")

        # 22개 제한 재적용
        matched_data = matched_data[:self.MAX_QUESTIONS]

        # 매칭 결과 로깅
        total_questions = len(questions)
        final_count = len(matched_data)

        logger.info(f"📊 매칭 완료:")
        logger.info(f"  - 전체 문제: {total_questions}개")
        logger.info(f"  - 최종 포함: {final_count}개")
        logger.info(f"  - 정답 매칭: {matched_count}개")
        logger.info(f"  - 정답 없음: {final_count - matched_count}개")

        return matched_data
    
    def _is_complete_question_data(self, question_data: Dict[str, Any]) -> bool:
        """
        문제 데이터가 완전한지 검증
        
        Args:
            question_data: 검증할 문제 데이터
            
        Returns:
            bool: 완전한 데이터 여부
        """
        required_fields = ["question_number", "content", "correct_answer"]
        
        for field in required_fields:
            value = question_data.get(field)
            if value is None or (isinstance(value, str) and not value.strip()):
                return False
        
        # 선택지가 있는 경우 검증
        options = question_data.get("options", {})
        if options and len(options) < 2:
            return False
        
        return True
    
    def convert_to_db_format(self, matched_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        매칭된 데이터를 데이터베이스 테이블 형식으로 변환
        """
        questions = []
        answer_options = []
        correct_answers = []
        
        current_time = datetime.now(timezone.utc).isoformat()
        
        for idx, item in enumerate(matched_data, 1):
            question_id = idx
            
            # Question 레코드
            question_record = {
                "id": question_id,
                "content": item.get("content", ""),
                "description": item.get("description", None),
                "difficulty": item.get("difficulty", "중"),
                "subject": item.get("subject", ""),
                "area_name": item.get("area_name", ""),
                "is_active": True,
                "question_metadata": {
                    "question_number": item.get("question_number"),
                    "year": item.get("year"),
                    "answer_source": item.get("answer_source", ""),
                },
                "created_at": current_time,
                "updated_at": current_time
            }
            
            # 간단 버전 필드 추가
            options = item.get("options", {})
            if options:
                question_record["choices"] = list(options.values())
                question_record["correct_answer"] = item.get("correct_answer", "")
            
            questions.append(question_record)
            
            # AnswerOption 레코드들
            for option_label, option_text in options.items():
                option_record = {
                    "question_id": question_id,
                    "option_text": option_text,
                    "option_label": option_label,
                    "display_order": int(option_label) if option_label.isdigit() else 0,
                    "created_at": current_time,
                    "updated_at": current_time
                }
                answer_options.append(option_record)
            
            # CorrectAnswer 레코드
            correct_answer = item.get("correct_answer")
            if correct_answer:
                correct_answer_record = {
                    "question_id": question_id,
                    "answer_text": correct_answer,
                    "created_at": current_time,
                    "updated_at": current_time
                }
                correct_answers.append(correct_answer_record)
        
        logger.info(f"DB 형식 변환 완료: Questions({len(questions)}), Options({len(answer_options)}), Answers({len(correct_answers)})")
        
        return {
            "questions": questions,
            "answer_options": answer_options,
            "correct_answers": correct_answers
        }

    # ============= 통합 AI 응답 처리 유틸리티들 =============
    
    @staticmethod
    def parse_ai_json_response(response_text: str, fallback_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        AI 응답에서 JSON 추출 및 파싱 (모든 AI 서비스에서 공통 사용)
        
        Args:
            response_text: AI 응답 텍스트
            fallback_data: 파싱 실패시 기본값
            
        Returns:
            파싱된 JSON 데이터
        """
        try:
            # 1단계: JSON 블록 추출 시도
            import re
            
            # ```json ... ``` 블록 찾기
            json_block_match = re.search(r'```json\s*(.*?)\s*```', response_text, re.DOTALL)
            if json_block_match:
                json_str = json_block_match.group(1).strip()
            else:
                # { ... } 패턴 찾기
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group().strip()
                else:
                    # [ ... ] 배열 패턴 찾기
                    array_match = re.search(r'\[.*\]', response_text, re.DOTALL)
                    if array_match:
                        json_str = array_match.group().strip()
                    else:
                        json_str = response_text.strip()
            
            # 2단계: JSON 파싱 시도
            cleaned_json = QuestionParser._clean_json_text_unified(json_str)
            result = json.loads(cleaned_json)
            
            logger.debug(f"✅ AI JSON 파싱 성공: {len(str(result))} 문자")
            return result
            
        except json.JSONDecodeError as e:
            logger.warning(f"⚠️ AI JSON 파싱 실패, 적극적 정리 시도: {e}")
            
            try:
                # 3단계: 적극적 JSON 정리
                aggressive_cleaned = QuestionParser._aggressive_json_clean_unified(response_text)
                result = json.loads(aggressive_cleaned)
                
                logger.info("✅ 적극적 JSON 정리로 파싱 성공")
                return result
                
            except json.JSONDecodeError as e2:
                logger.error(f"❌ 모든 JSON 파싱 시도 실패: {e2}")
                
                # 4단계: 폴백 데이터 반환
                if fallback_data:
                    logger.info("📋 폴백 데이터 사용")
                    return fallback_data
                else:
                    logger.info("📋 기본 에러 구조 반환")
                    return {
                        "error": "JSON 파싱 실패",
                        "raw_response": response_text[:200] + "..." if len(response_text) > 200 else response_text,
                        "parse_attempted": True
                    }
        
        except Exception as e:
            logger.error(f"❌ AI 응답 처리 중 예외 발생: {e}")
            return fallback_data or {"error": str(e), "parse_attempted": True}
    
    @staticmethod
    def _clean_json_text_unified(text: str) -> str:
        """통합 JSON 텍스트 정리 (모든 서비스 공통 사용)"""
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # // 주석 제거
            if '//' in line:
                # JSON 문자열 내부의 //는 보존
                in_string = False
                escaped = False
                cleaned_line = ""
                
                for i, char in enumerate(line):
                    if escaped:
                        cleaned_line += char
                        escaped = False
                        continue
                    
                    if char == '\\':
                        escaped = True
                        cleaned_line += char
                        continue
                    
                    if char == '"' and not escaped:
                        in_string = not in_string
                        cleaned_line += char
                        continue
                    
                    if not in_string and char == '/' and i + 1 < len(line) and line[i + 1] == '/':
                        # 주석 시작, 나머지 줄 무시
                        break
                    
                    cleaned_line += char
                
                line = cleaned_line
            
            # /* */ 주석 제거 (단순 버전)
            while '/*' in line and '*/' in line:
                start = line.find('/*')
                end = line.find('*/', start) + 2
                line = line[:start] + line[end:]
            
            # 빈 줄이 아니면 추가
            if line.strip():
                cleaned_lines.append(line)
        
        return '\n'.join(cleaned_lines)
    
    @staticmethod
    def _aggressive_json_clean_unified(text: str) -> str:
        """적극적 JSON 정리 (모든 서비스 공통 사용)"""
        import re
        
        # 마지막 } 또는 ] 이후의 모든 텍스트 제거
        text = text.strip()
        
        # JSON 배열인지 객체인지 확인
        if text.startswith('['):
            # 마지막 ]의 위치 찾기
            last_bracket = text.rfind(']')
            if last_bracket != -1:
                text = text[:last_bracket + 1]
        elif text.startswith('{'):
            # 마지막 }의 위치 찾기
            last_brace = text.rfind('}')
            if last_brace != -1:
                text = text[:last_brace + 1]
        
        # 불완전한 JSON 키-값 수정
        text = re.sub(r',\s*}', '}', text)  # 마지막 콤마 제거
        text = re.sub(r',\s*]', ']', text)  # 배열 마지막 콤마 제거
        text = re.sub(r'([^"])\s*:\s*([^"\[\{].*?)([,\}\]])', r'\1: "\3"\3', text)  # 값 따옴표 추가
        
        return text
    
    @staticmethod 
    def extract_ai_content_patterns(response_text: str, patterns: Dict[str, str]) -> Dict[str, str]:
        """
        AI 응답에서 특정 패턴 추출 (모든 AI 서비스 공통)
        
        Args:
            response_text: AI 응답 텍스트
            patterns: {"key": "regex_pattern"} 형태의 추출 패턴들
            
        Returns:
            추출된 데이터 딕셔너리
        """
        import re
        
        extracted = {}
        
        for key, pattern in patterns.items():
            try:
                match = re.search(pattern, response_text, re.IGNORECASE | re.DOTALL)
                if match:
                    if match.groups():
                        extracted[key] = match.group(1).strip()
                    else:
                        extracted[key] = match.group(0).strip()
                else:
                    extracted[key] = ""
                    
            except Exception as e:
                logger.warning(f"패턴 추출 실패 ({key}): {e}")
                extracted[key] = ""
        
        return extracted
    
    @staticmethod
    def validate_ai_analysis_result(
        analysis_result: Dict[str, Any], 
        required_fields: List[str],
        default_values: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """
        AI 분석 결과 검증 및 보완 (모든 AI 분석 서비스 공통)
        
        Args:
            analysis_result: AI 분석 결과
            required_fields: 필수 필드 목록
            default_values: 기본값 딕셔너리
            
        Returns:
            검증된 분석 결과
        """
        validated = analysis_result.copy() if analysis_result else {}
        defaults = default_values or {}
        
        # 필수 필드 확인 및 기본값 설정
        for field in required_fields:
            if field not in validated or not validated[field]:
                if field in defaults:
                    validated[field] = defaults[field]
                    logger.debug(f"📋 기본값 적용: {field} = {defaults[field]}")
                else:
                    validated[field] = None
                    logger.warning(f"⚠️ 필수 필드 누락: {field}")
        
        # 메타데이터 추가
        validated["validation_timestamp"] = datetime.now().isoformat()
        validated["validation_applied"] = True
        
        return validated

# 싱글톤 인스턴스
question_parser = QuestionParser()